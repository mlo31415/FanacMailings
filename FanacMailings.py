from __future__ import annotations

from dataclasses import dataclass, field
import csv
import os
import re
import datetime

import openpyxl

from FanzineIssueSpecPackage import FanzineDate
from Settings import Settings
from HelpersPackage import FindAndReplaceBracketedText, ParseFirstStringBracketedText, SortMessyNumber, SortTitle, Pluralize, NormalizePersonsName, Int0
from HelpersPackage import FindIndexOfStringInList, FormatCount, DebuggerIsRunning, UnicodeToHtml, MakeFancyLink
from Log import LogError, Log, LogDisplayErrorsIfAny, LogOpen


def main():
    LogOpen("log.txt", "log-ERRORS.txt")
    if not Settings().Load("FanacMailings settings.txt", MustExist=True, SuppressMessageBox=True):
        LogError("Could not find settings file 'FanacMailings settings.txt'")
        return

    # **************************************************************************
    # Get the list of known apas
    knownApas=Settings().Get("Known APAs")
    if len(knownApas) == 0:
        LogError("Settings file 'FanacMailings settings.txt' does not contain a value for 'Known APAs' (the list of APAs we care about here)")
        return
    knownApas=[x.replace('"', '').strip() for x in knownApas.split(",")]

    # **************************************************************************
    # for each known apa, read Joe's APA mailings data if it exists
    # Mailings is a dictionary indexed by the apa name.
    #   The value is a dictionary indexed by the mailing number as a string
    #       The value of *that* is a MailingDev
    # Note that we do  not fill in Counts here
    mailingsInfoTablefromJoe: dict[str, dict[str, MailingInfoFromJoe]]={}
        # 1st level key is APA name
        # 2nd level key is mailing name
    for apaName in knownApas:
        table=ReadXLSX(apaName)
        if table is None:
            table={}
        mailingsInfoTablefromJoe[apaName]=table

    # **************************************************************************
    # Get the location of the CSV source file (generated by FanacAnalyzer) out of settings
    sourceCSVfile=Settings().Get("CSVSource")
    if len(sourceCSVfile) == 0:
        LogError("Settings file 'FanacMailings settings.txt' does not contain a value for CSVSource (the file generated by FanacAnalyzer)")
        return
    # Open and read it
    try:
        with open(sourceCSVfile, 'r') as csvfile:
            filereader=csv.reader(csvfile, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            mailingsdata=[x for x in filereader]
    except FileNotFoundError:
        LogError(f"Could not open CSV file {sourceCSVfile}")
        return

    if len(mailingsdata) < 100:
        LogError(f"There are {len(mailingsdata)} items in {sourceCSVfile} -- there should be many hundreds")
        if not DebuggerIsRunning():
            return

    # Segregate the headers info
    mailingsHeaders=mailingsdata[0]
    mailingsdata=mailingsdata[1:]

    # ---------------------------
    # Turn the data from FanacAnalyzer into a dictionary of the form dict(apa, dict(mailing, data)) by loading
    # the individual fanzine issue information read from the file from FanacAnalyzer
    # Allmailings is keyed by the apa's name.  The Value is an EntireAPA object
    allAPAs: AllAPAs=AllAPAs()
    for row in mailingsdata:
        fanzine=FanzineInMailing(mailingsHeaders, row)
        # The mailings column is of the form   ['FAPA 20 & VAPA 23']
        mailings=fanzine.Mailings.removeprefix("['").removesuffix("']")
        mailings=[x.strip() for x in mailings.split("&")]
        for mailing in mailings:
            for apaName in knownApas:
                m=re.match(rf"{apaName}\s(.*)$", mailing)
                if m is not None:
                    mailingNumber=m.groups()[0]
                    allAPAs[apaName][mailingNumber].append(fanzine)
                    break

    # ------------------
    # We've slurped in all the data.
    # Now merge Joe's mailing info into allAPAs
    for apa in allAPAs:
        for mailing in apa:
            if apa.Name in mailingsInfoTablefromJoe:
                if mailing.Name in mailingsInfoTablefromJoe[apa.Name]:
                    mailing.MIFJ=mailingsInfoTablefromJoe[apa.Name][mailing.Name]


    # The next step is to generate the counts
    # Walk through allAPAs
    # For each APA that we found there extract the data, merge it was Joe's data, and create a unified dataset to generate the web pages
    countAllAPAs=Counts()       # This is the only 'bare' Counts -- all the others are in larger structures
    for apa in allAPAs:

        # For each mailing of that APA count up the issues and pages
        for mailing in apa:
            # Sort the contributions into order by fanzine name
            #apa[mailingName]=sorted(apa[mailingName], key=lambda x: SortTitle(x.IssueName))

            # Now generate the data rows in the mailings table
            for apazine in mailing:
                mailing.Count+=Counts(Issues=1, Pages=apazine.PageCount)
            apa.Count+=Counts(Mailings=1, Issues=mailing.Count.Issues, Pages=mailing.Count.Pages)

        countAllAPAs+=Counts(Mailings=apa.Count.Mailings, Issues=apa.Count.Issues, Pages=apa.Count.Pages)


    ##################################################################################################################
    # We have done all the analysis: generate the HTML pages

    # We will create a file in the ReportsDir for each APA, and put the individual issue index pages there
    reportsdir=Settings().Get("ReportsDir")
    if len(reportsdir) == 0:
        LogError("Settings file 'FanacMailings settings.txt' does not contain a value for ReportsDir (the directory to be used for reports)")
        return
    if not os.path.exists(reportsdir):
        os.mkdir(reportsdir)

    # All the pages we generate here need the same kinds of information to be added:
    #   Page title
    #   Page metadata
    #   Updated timestamp
    def AddBoilerplate(page: str, title: str, metadata: str) -> str:
        start, mid, end=ParseFirstStringBracketedText(page, "fanac-title")
        mid=mid.replace("title of page", title)
        page=start+mid+end

        start, mid, end=ParseFirstStringBracketedText(page, "head")
        mid=mid.replace("mailing content", metadata)
        page=start+mid+end

        # Add the updated date/time
        page, _=FindAndReplaceBracketedText(page, "fanac-updated", f"Updated {datetime.datetime.now().strftime('%m/%d/%Y, %H:%M:%S')}")
        return page


    # Read the mailing template file
    templateFilename=Settings().Get("Template-Mailing")
    if len(templateFilename) == 0:
        LogError("Settings file 'FanacMailings settings.txt' does not contain a value for Template-Mailing (the name of the template file for an individual mailing page)")
        return
    try:
        with open(templateFilename, "r") as file:
            templateMailing="".join(file.readlines())
    except FileNotFoundError:
        LogError(f"Could not open the mailing template file: '{templateFilename}'")
        return

    # Read the apa template file
    templateFilename=Settings().Get("Template-APA")
    if len(templateFilename) == 0:
        LogError("Settings file 'FanacMailings settings.txt' does not contain a value for Template-APA (the template for an APA page)")
        return
    try:
        with open(templateFilename, "r") as file:
            templateApa="".join(file.readlines())
    except FileNotFoundError:
        LogError(f"Could not open the APA template file, '{templateFilename}'")
        return

    # Walk through the info generated by FanacAnalyzer.
    # For each APA that we found there:
    #   Create an apa HTML page listing (and linking to) all the mailing pages
    #   Create all the individual mailing pages
    allAPAs.sort()
    for apa in allAPAs:


        # Make sure that a directory exists for this APA's html files
        if not os.path.exists(os.path.join(reportsdir, apa.Name)):
            os.mkdir(os.path.join(reportsdir, apa.Name))

        apa.sort()
        for mailing in apa:
            mailing.sort()

            ##################################################################
            ##################################################################
            # Do a mailing page
            # First, the top matter
            # <div><fanac-top>
            # <table class=topmatter>
            # <tr><td class=topmatter>mailing</td></tr>
            # <tr><td class=topmatter>editor</td></tr>
            # <tr><td class=topmatter>date</td></tr>
            # </table>
            # </fanac-top></div>
            mailingPage=templateMailing
            start, mid, end=ParseFirstStringBracketedText(mailingPage, "fanac-top")
            editor=f"OE: {NormalizePersonsName(mailing.MIFJ.Editor)}"
            when=mailing.MIFJ.Date.FormatDate("%B %Y")
            number=mailing.Name
            mid=mid.replace("editor", editor)
            mid=mid.replace("date", when)
            mid=mid.replace("mailing", f"{apa.Name} Mailing #{number}")
            mailingPage=start+mid+end

            mailingPage=AddBoilerplate(mailingPage, f"{apa.Name}-{mailing.Name}", f"{mailing.Name}, {editor}, {when}, {apa.Name}-mailing")

            # Now the bottom matter (the list of fanzines)
            newtable="<tr>\n"
            # Generate the header row, selecting only those headers which are in this dict:
            colSelectionAndOrder=["IssueName", "Editor", "PageCount"]   # The columns to be displayed in order

            newtable+="<th>Contribution</th>\n"
            newtable+="<th>Editor</th>\n"
            newtable+="<th>Pages</th>\n"
            newtable+="</tr>\n"

            # Now generate the data rows in the mailings table
            for apazine in mailing:
                countThisIssue=Counts()
                countThisIssue.Issues=1
                newtable+="<tr>\n"
                if apazine.DirURL != "" and apazine.PageName != "":
                    href=f"{apazine.DirURL}/{apazine.PageName}>"
                    href=href.replace(" ", "%20")
                    newtable+=f"<td><a href={href}{UnicodeToHtml(apazine.IssueName)}</a></td>\n"
                else:
                    newtable+=f"<td>&nbsp;</td>\n"
                if apazine.Editor != "":
                    newtable+=f"<td>{MakeFancyLink(apazine.Editor)}&nbsp;&nbsp;</td>"
                else:
                    newtable+=f"<td>&nbsp;</td>\n"
                if apazine.PageCount != "":
                    newtable+=f"<td>{apazine.PageCount}</td>\n"
                    countThisIssue+=Int0(apazine.PageCount)
                else:
                    newtable+=f"<td>&nbsp;</td>\n"
                newtable+="</tr>\n"
            newtable=newtable.replace("\\", "/")

            # Insert the new issues table into the template
            mailingPage, success=FindAndReplaceBracketedText(mailingPage, "fanac-rows", newtable)
            if not success:
                LogError(f"Could not add issues table to mailing page {templateFilename} at 'fanac-rows'")
                return

            # Insert the label for the button taking you to the previous mailing for this APA
            index=apa.prevIndex(mailing.Name)
            if index is None:
                buttonText=f"No prev mailing "
                link=""
            else:
                buttonText=f" Prev Mailing (#{apa[index].Name}) "
                link=f'"{apa[index].Name}.html"'
            mailingPage, success=FindAndReplaceBracketedText(mailingPage, "fanac-PrevMailing", buttonText)
            if success:
                mailingPage=mailingPage.replace('"prev.html"', link)
            if not success:
                LogError(f"Could not change prev button text on mailing page {templateFilename} at 'fanac-PrevMailing'")
                return

            # Insert the label for the button taking you up one level to all mailings for this APA
            mailingPage, success=FindAndReplaceBracketedText(mailingPage, "fanac-AllMailings", f"All {apa.Name} mailings")
            if not success:
                LogError(f"Could not change up to APA button text on mailing page {templateFilename} at 'fanac-AllMailings'")
                return

            # Insert the label for the button taking you to the next mailing for this APA
            index=apa.nextIndex(mailing.Name)
            if index is None:
                buttonText=f"No next mailing "
                link=""
            else:
                buttonText=f" Next Mailing (#{apa[index].Name}) "
                link=f'"{apa[index].Name}.html"'
            mailingPage, success=FindAndReplaceBracketedText(mailingPage, "fanac-NextMailing", buttonText)
            if success:
                mailingPage=mailingPage.replace('"next.html"', link)
            if not success:
                LogError(f"Could not change next button text on mailing page {templateFilename} at 'fanac-NextMailing'")
                return

            # Modify the Mailto: so that the page name appears as the subject
            mailingPage, success=FindAndReplaceBracketedText(mailingPage, "fanac-ThisPageName", f"{apa.Name}:{mailing.Name}")
            if not success:
                LogError(f"Could not change mailto Subject on mailing page {templateFilename} at 'fanac-ThisPageName'")
                #return

            # Add counts of mailings and contributions to bottom
            start, mid, end=ParseFirstStringBracketedText(mailingPage, "fanac-totals")
            mailingPage=f"{start} {mailing.Count}  {end}"

            # Write the mailing file
            with open(os.path.join(reportsdir, apa.Name, mailing.Name)+".html", "w") as file:
                mailingPage=mailingPage.split("\n")
                file.writelines(mailingPage)


        ##################################################################
        ##################################################################
        # Now that the mailing pages are all done, do an apa page

        # Add the APA's name at the top
        start, mid, end=ParseFirstStringBracketedText(templateApa, "fanac-top")
        mid=mid.replace("apa-name", apa.Name)
        newAPAPage=start+mid+end

        # Add random descriptive information if a file <apa>-bumpf.txt exists.  (E.g., SAPS-bumpf.txt)
        fname=apa.Name+"-bumpf.txt"
        if os.path.exists(fname):
            with open(fname, "r") as file:
                bumpf=file.read()
            if len(bumpf) > 0:
                start, mid, end=ParseFirstStringBracketedText(newAPAPage, "fanac-bumpf")
                if len(end) > 0:
                    mid=bumpf+"<p>"
                    newAPAPage=start+mid+end
            Log(f"Bumpf added to {apa.Name} page")
        else:
            Log(f" No {fname} file found, so no bumpf added to {apa.Name} page.")

        newAPAPage=AddBoilerplate(newAPAPage, f"{apa.Name} Mailings", f"{apa.Name} mailings")

        loc=newAPAPage.find("</fanac-rows>")
        if loc < 0:
            LogError(f"The APA template '{templateFilename}' is missing the '</fanac-rows>' indicator.")
            return
        newAPAPageFront=newAPAPage[:loc]
        newAPAPageRear=newAPAPage[loc+len("</fanac-rows>"):]

        apa.sort()
        for mailing in apa:
            when=mailing.MIFJ.Date
            editor=mailing.MIFJ.Editor
            issues=mailing.Count.Issues
            pages=mailing.Count.Pages
            newAPAPageFront+=(f"\n<tr><td><a href={mailing.Name}.html>{mailing.Name}</a></td>"
                              f"<td>{when}</td><td>{editor}</td>"
                              f"<td style='text-align: right'>{issues}&nbsp;&nbsp;&nbsp;&nbsp;</td>"
                              f"<td style='text-align: right'>{pages}&nbsp;&nbsp;&nbsp;&nbsp;</td>"
                              f"</tr>")

        newAPAPage=newAPAPageFront+newAPAPageRear

        # Add counts of mailings and contributions to bottom
        start, mid, end=ParseFirstStringBracketedText(newAPAPage, "fanac-totals")
        newAPAPage=f"{start} {apa.Count}  {end}"

        # Add the updated date/time
        newAPAPage, success=FindAndReplaceBracketedText(newAPAPage, "fanac-updated", f"Updated {datetime.datetime.now().strftime('%m/%d/%Y, %H:%M:%S')}>")

        # Make the mailto correctly list the apa in the subject line
        newAPAPage, success=FindAndReplaceBracketedText(newAPAPage, "fanac-APAPageMailto", f"Issue related to APA {apa.Name}")
        if not success:
            LogError(f"The APA template '{templateFilename}' is missing the '</fanac-APAPageMailto>' indicator.")
            return

        # Write out the APA list of all mailings
        with open(os.path.join(reportsdir, apa.Name, "index.html"), "w") as file:
            file.writelines(newAPAPage)

    ##################################################################
    ##################################################################
    # Generate the All Apas root page

    templateFilename=Settings().Get("Template-allAPAs")
    if len(templateFilename) == 0:
        LogError("Settings file 'FanacMailings settings.txt' does not contain a value for Template-allAPAs (the template for the page listing all APAs)")
        return
    try:
        with open(templateFilename, "r") as file:
            templateAllApas="".join(file.readlines())
    except FileNotFoundError:
        LogError(f"Could not open the all APAs template file, '{templateFilename}'")
        return

    templateAllApas=AddBoilerplate(templateAllApas, f"Mailings for All APAs", f"Mailings for All APAs")

    listText="\n<i>Click on the APA's name to see APA's contents</i>\n"
    listText+="<style>th, td{border-style: hidden;}</style>\n\n"

    listText+="<table>\n<tr>\n<th>&nbsp;&nbsp;&nbsp;APA</th>\n<th>&nbsp;Mailings&nbsp;</th>\n<th>&nbsp;Apazines&nbsp;</th>\n<th>&nbsp;Pages&nbsp;</th</tr>\n"

    allAPAs.sort()
    for apa in allAPAs:
        listText+=(f"\n<tr><td>&nbsp;&nbsp;&nbsp;<a href={apa.Name}/index.html>{apa.Name}</a></td>\n"
                          f"<td style='text-align: right'>{apa.Count.Mailings}&nbsp;&nbsp;&nbsp;</td>\n"
                          f"<td style='text-align: right'>{apa.Count.Issues}&nbsp;&nbsp;&nbsp;</td>\n"
                          f"<td style='text-align: right'>{FormatCount(apa.Count.Pages)}&nbsp;&nbsp;&nbsp;</td>\n"
                          f"</tr>\n")
    # Add counts of mailings and contributions to bottom
    for apa in allAPAs:
        allAPAs.Count+=apa.Count
    listText+=(f"\n<tr><td>&nbsp;&nbsp;&nbsp;&nbsp</td>\n"
               f"<td style='text-align: right'>______&nbsp;&nbsp;</td>\n"
               f"<td style='text-align: right'>______&nbsp;&nbsp;</td>\n"
               f"<td style='text-align: right'>______&nbsp;&nbsp;</td>\n")
    listText+=(f"\n<tr><td>&nbsp;&nbsp;&nbsp;&nbsp</td>\n"
               f"<td style='text-align: right'>{allAPAs.Count.Mailings}&nbsp;&nbsp;&nbsp;</td>\n"
               f"<td style='text-align: right'>{allAPAs.Count.Issues}&nbsp;&nbsp;&nbsp;</td>\n"
               f"<td style='text-align: right'>{FormatCount(allAPAs.Count.Pages)}&nbsp;&nbsp;&nbsp;</td>\n"
               f"</tr>\n")

    listText+="</table>\n"
    templateAllApas, success=FindAndReplaceBracketedText(templateAllApas, "fanac-list", listText)

    with open(os.path.join(reportsdir, "index.html"), "w") as file:
        file.writelines(templateAllApas)

# End Main
###################################################################


# Read the APA Mailings.xlsx file supplied by Joe to get OE, date, etc., information for each mailing.
def ReadXLSX(apaName: str) -> dict[str, MailingInfoFromJoe] | None:
    xlsxname="APA Mailings.xlsx"
    # Skip missing xlsx files
    if not os.path.exists(xlsxname):
        LogError("Can't find {xlsxname}")
        return None
    # Read the apa mailings file
    try:
        wb=openpyxl.load_workbook(filename=xlsxname)
    except FileNotFoundError:
        LogError(f"Could not open xlsx file {xlsxname}")
        return None


    if apaName not in wb.sheetnames:
        return None
    ws=wb[apaName]

    # Separate out the header row
    mailingsheaders=[x.value for x in ws[1]]

    monthCol=FindIndexOfStringInList(mailingsheaders, "Month")
    if monthCol is None:
        LogError(f"{xlsxname} does not contain a 'Month' column")
        return None
    yearCol=FindIndexOfStringInList(mailingsheaders, "Year")
    if yearCol is None:
        LogError(f"{xlsxname} does not contain a 'Year' column")
        return None
    editorCol=FindIndexOfStringInList(mailingsheaders, ["Editor", "OE"])
    if editorCol is None:
        LogError(f"{xlsxname} does not contain an 'Editor' or an 'OE' column")
        return None
    mailingCol=FindIndexOfStringInList(mailingsheaders, ["Mailing", "Issue"])
    if mailingCol is None:
        LogError(f"{xlsxname} does not contain a 'Mailing' or an 'Issue' column")
        return None

    mailingsInfoFromJoe={}
    for i in range(2, 10000):
        row=[x.value for x in ws[i]]
        if all([x is None for x in row]):
            break
        mailingNum=row[mailingCol]
        if type(mailingNum) is int:
            mailingNum=str(mailingNum)  # Standard is to treat mailing number as a string, because sometimes it has to be
        editor=row[editorCol]
        if editor is None:
            editor=""
        mailingsInfoFromJoe[mailingNum]=MailingInfoFromJoe(Number=mailingNum, Year=row[yearCol], Month=row[monthCol], Editor=editor)
    return mailingsInfoFromJoe


######################################################################
# A class to count mailings, issues and pages
class Counts:
    def __init__(self, Pages: int|str=0, Issues: int=0, Mailings: int=0):
        self.Mailings=Mailings
        self.Issues=Issues
        if type(Pages) is str:
            Pages=Int0(Pages)
        self.Pages=Pages

    def __hash__(self):
        return self.Mailings.__hash__()+self.Pages.__hash__()+self.Issues.__hash__()

    def __iadd__(self, val:Counts | int):
        self.Add(val)
        return self

    def __str__(self):
        s=""
        if self.Mailings > 0:
            s+=f"{Pluralize(self.Mailings, 'mailing')}, "
        return s+f"{Pluralize(self.Issues, 'issue')}, {Pluralize(self.Pages, 'page')}"

    # Add a Count or a single fanzine
    def __add__(self, val:Counts | int) -> Counts:
        temp=Counts(Pages=self.Pages, Issues=self.Issues, Mailings=self.Mailings)
        temp.Add(val)
        return temp

    def Add(self, val:Counts | int):
        if type(val) is Counts:
            self.Mailings+=val.Mailings
            self.Issues+=val.Issues
            self.Pages+=val.Pages
            return
        if type(val) is int:
            if self.Mailings == 0:
                self.Mailings=1
            if self.Issues == 0:
                self.Issues=1
            self.Pages+=val
            return
        assert False


######################################################################
# Entry for a specific mailing in a dictionary of mailings for an APA.
class MailingInfoFromJoe:
    def __init__(self, Number: str = "", Year: str = "", Month: str = "", Editor: str = ""):
        self.Number: str=Number
        self.Editor: str=Editor
        self.Prev: str=""
        self.Next: str=""

        fd=FanzineDate()
        if Month != "":
            fd.Month=Month
        if Year != "":
            fd.Year=Year
        self.Date: FanzineDate=fd

    def __hash__(self):
        return self.Number.__hash__()+self.Editor.__hash__()+self.Prev.__hash__()+self.Next.__hash__()+self.Date.__hash__()


    @property
    def Year(self) -> int:
        return self.Date.Year

    @Year.setter
    def Year(self, y: int) -> None:
        self.Date.Year=y

    @property
    def Month(self) -> int:
        return self.Date.Month

    @Month.setter
    def Month(self, m: int) -> None:
        self.Date.Month=m
# --- end class MailingDev ---


######################################################################
#
class OneMailing:
    def __init__(self):
        self._Count: Counts=Counts()      # The totals for all the apazines in the mailing
        self.MIFJ: MailingInfoFromJoe=MailingInfoFromJoe()       # Joe's info on the mailing
        self.ListFIM: list=[]        # A list of all the apazines in the mailing
        self.Name: str=""        # The name of the mailing (usually a number.)

    def append(self, val: FanzineInMailing):
        self.ListFIM.append(val)
    def __len__(self):
        return len(self.ListFIM)
    def __hash__(self):
        h=self.Name.__hash__()+self.Count.__hash__()+self.MIFJ.__hash__()
        for lf in self.ListFIM:
            h+=lf.__hash__()
        return h

    def __iter__(self):
        self._current=0
        return self

    def __next__(self):
        if self._current >= len(self.ListFIM):
            raise StopIteration
        self._current += 1
        return self.ListFIM[self._current-1]

    def sort(self):
        self.ListFIM.sort(key=lambda x: SortTitle(x.IssueName))

    @property
    def Count(self):
        return self._Count
    @Count.setter
    def Count(self, val):
        self._Count=val



@dataclass
class EntireAPA:
    Count: Counts=field(default_factory=lambda: Counts())
    List: list[OneMailing]=field(default_factory=list)
    Name: str=""

    def __hash__(self):
        h=0
        for om in self.List:
            h+=om.__hash__()
        return h+self.Name.__hash__()+self.Count.__hash__()

    def __len__(self) -> int:
        return len(self.List)

    def append(self, val:OneMailing):
        self.List.append(val)
    def __getitem__(self, index: str) -> OneMailing:
        for (i, x) in enumerate(self.List):
            if x.Name == index:
                return x
        new=OneMailing()
        new.Name=index
        self.List.append(new)
        return new

    def nextIndex(self, index: str) -> str | None:
        for (i, x) in enumerate(self.List):
            if x.Name == index:
                if i+1 >= len(self.List):
                    return None
                return self.List[i+1].Name
        return None

    def prevIndex(self, index: str) -> str|None:
        for (i, x) in enumerate(self.List):
            if x.Name == index:
                if i-1 < 0:
                    return None
                return self.List[i-1].Name
        return None

    def __iter__(self):
        self._current=0
        return self

    def __next__(self):
        if self._current >= len(self.List):
            raise StopIteration
        self._current += 1
        return self.List[self._current-1]

    def sort(self):
        self.List.sort(key=lambda x: SortMessyNumber(x.Name))


@dataclass
class AllAPAs:
    Count: Counts=field(default_factory=lambda: Counts())
    List: list[EntireAPA]=field(default_factory=list)

    def append(self, val:EntireAPA):
        self.List.append(val)

    def __getitem__(self, index: str) -> EntireAPA:
        for (i, x) in enumerate(self.List):
            if x.Name == index:
                return x
        new=EntireAPA()
        new.Name=index
        self.List.append(new)
        return new

    def __iter__(self):
        self._current=0
        return self

    def __next__(self):
        if self._current >= len(self.List):
            raise StopIteration

        self._current += 1
        return self.List[self._current-1]

    def sort(self):
        self.List.sort(key=lambda x: x.Name)


######################################################################
# A class to hold the information for one fanzine in one mailing of an APA

class FanzineInMailing:
    def __init__(self, headers: list[str], row: list[str]):
        self.IssueName: str=self.initialize(headers, row, "IssueName")
        self.Series: str=self.initialize(headers, row, "Series")
        self.SeriesName: str=self.initialize(headers, row, "SeriesName")
        self.DisplayName: str=self.initialize(headers, row, "DisplayName")
        self.DirURL: str=self.initialize(headers, row, "DirURL")
        self.PageName: str=self.initialize(headers, row, "PageName")
        self.FIS: str=self.initialize(headers, row, "FIS")
        self.Locale: str=self.initialize(headers, row, "Locale")
        self.PageCount: str=self.initialize(headers, row, "PageCount")
        self.Editor: str=self.initialize(headers, row, "Editor")
        self.TagList: str=self.initialize(headers, row, "TagList")
        self.Mailings: str=self.initialize(headers, row, "Mailings")


    @staticmethod
    def initialize(headers: list[str], row: list[str], item: str) -> str:
        index=FindIndexOfStringInList(headers, item)
        if index == -1:
            return ""
        return row[index]

# --- end class FanzineInMailing ---


# Run main()
if __name__ == "__main__":
    main()
    LogDisplayErrorsIfAny()


